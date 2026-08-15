"""Export Excel (.xlsx) — réunion unique ou ensemble des réunions."""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Point

NAVY_FILL = PatternFill("solid", fgColor="0B2545")
NAVY_FONT = Font(color="FFFFFF", bold=True, name="Calibri")
RED_FILL = PatternFill("solid", fgColor="C0392B")
RED_FONT = Font(color="FFFFFF", bold=True)
THIN = Border(
    left=Side(style="thin", color="9AA8B8"),
    right=Side(style="thin", color="9AA8B8"),
    top=Side(style="thin", color="9AA8B8"),
    bottom=Side(style="thin", color="9AA8B8"),
)


def _style_header(ws, ncol):
    for col in range(1, ncol + 1):
        cell = ws.cell(1, col)
        cell.fill = NAVY_FILL
        cell.font = NAVY_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{ws.max_row}"
    ws.freeze_panes = "A2"


def _autosize(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 10
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            length = max(length, min(len(val) + 2, 45))
        ws.column_dimensions[letter].width = length


def _urgency_cf(ws, col_letter, start_row, end_row):
    if end_row < start_row:
        return
    rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=["5"], fill=RED_FILL, font=RED_FONT),
    )


def _write_points(ws, points, with_reunion=False):
    headers = []
    if with_reunion:
        headers.append("Réunion")
    headers += [
        "N°",
        "Rubrique",
        "Service",
        "Sujet",
        "Décision",
        "Action",
        "Responsable",
        "Délai",
        "Urgence",
        "Statut",
    ]
    ws.append(headers)
    for point in points:
        row = []
        if with_reunion:
            row.append(f"{point.reunion.date:%d/%m/%Y} — {point.reunion.get_type_display()}")
        row += [
            point.numero,
            point.get_rubrique_display(),
            point.get_service_display(),
            point.sujet,
            point.decision,
            point.action,
            point.responsable,
            point.delai.strftime("%d/%m/%Y") if point.delai else "",
            point.urgence,
            point.get_statut_display(),
        ]
        ws.append(row)
    ncol = len(headers)
    _style_header(ws, ncol)
    urg_col = get_column_letter(ncol - 1)
    _urgency_cf(ws, urg_col, 2, ws.max_row)
    _autosize(ws)
    ws.sheet_properties.tabColor = "0B2545"


def export_reunion_xlsx(reunion):
    wb = Workbook()
    ws = wb.active
    ws.title = "Points"
    _write_points(ws, reunion.points.all(), with_reunion=False)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_all_reunions_xlsx(reunions_qs):
    reunions = list(reunions_qs.prefetch_related("points"))
    wb = Workbook()

    ws_r = wb.active
    ws_r.title = "Réunions"
    ws_r.append(["Date", "Type", "Président", "Nb participants", "Lieu", "Rapporteur"])
    for r in reunions:
        ws_r.append(
            [
                r.date.strftime("%d/%m/%Y"),
                r.get_type_display(),
                r.president,
                r.nombre_participants,
                r.lieu,
                r.rapporteur,
            ]
        )
    _style_header(ws_r, 6)
    _autosize(ws_r)

    ws_p = wb.create_sheet("Points")
    all_points = []
    for r in reunions:
        all_points.extend(list(r.points.all()))
    _write_points(ws_p, all_points, with_reunion=True)

    ws_s = wb.create_sheet("Statistiques")
    points = all_points
    ws_s.append(["Indicateur", "Valeur"])
    ws_s.append(["Nombre de réunions", len(reunions)])
    ws_s.append(["Nombre de points", len(points)])
    ws_s.append(["Points critiques (urgence 5)", sum(1 for p in points if p.urgence == 5)])
    ws_s.append(["Points urgents (4 et 5)", sum(1 for p in points if p.urgence >= 4)])
    ws_s.append(["Points à faire", sum(1 for p in points if p.statut == Point.STATUT_A_FAIRE)])
    ws_s.append(["Points en cours", sum(1 for p in points if p.statut == Point.STATUT_EN_COURS)])
    ws_s.append(["Points faits", sum(1 for p in points if p.statut == Point.STATUT_FAIT)])
    ws_s.append(["Points reportés", sum(1 for p in points if p.statut == Point.STATUT_REPORTE)])
    ws_s.append([])
    ws_s.append(["Points par service", "Nombre"])
    for code, label in Point.SERVICE_CHOICES:
        ws_s.append([label, sum(1 for p in points if p.service == code)])
    ws_s.append([])
    ws_s.append(["Urgence", "Nombre"])
    for level in range(1, 6):
        ws_s.append([level, sum(1 for p in points if p.urgence == level)])
    _style_header(ws_s, 2)
    _autosize(ws_s)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
